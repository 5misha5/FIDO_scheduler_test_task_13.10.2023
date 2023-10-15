import openpyxl
import re
from abc import ABC, abstractmethod

import zipfile
import xml.etree.ElementTree

from win32com import client as wc
import os


COLS = 6

DAYS_OF_WEEKS = 0
TIME = 1
COURSE = 2
GROUPS = 3
WEEKS = 4
LECT_HALL = 5


class Reader(ABC):
    """
    Abstract base class for data readers.

    Parameters:
        path (str): The path to the data file.

    """

    def __init__(self, path) -> None:
        """
        Initialize a Reader instance.

        Parameters:
            path (str): The path to the data file.

        """
        self.path = path

    @abstractmethod
    def read(self):
        """
        Abstract method for reading data from a file.

        Returns:
            list: The read schedule data.
        """
        pass

class XLSXReader(Reader):
    """
    A class for reading schedule data from XLSX files.

    Parameters:
        path (str): The path to the XLSX file.

    """
    def __init__(self, path) -> None:
        """
        Initialize an XLSXReader instance.

        Parameters:
            path (str): The path to the XLSX file.

        """
        super().__init__(path)
    
    def read(self):
        """
        Read data from an XLSX file.

        Returns:
            list: The read schedule data in a structured format.
        """
        path = self.path       
        
        ws = openpyxl.load_workbook(path).active

        schedule = []

        rows_range = self.get_schedule_rows_range(ws)

        for row in ws.iter_rows(rows_range[0], rows_range[1]):

            row_data = []

            for col in range(0,6):
                if col in {0,1}:
                    row_data.append(self.get_nearest_up_cell_val(row[col]))
                else:
                    cell_val = self.get_cell_val(row[col])

                    row_data.append(str(self.get_cell_val(row[col])))

            schedule.append(row_data)

        return schedule

    def get_cell_val(self, cell, default = None):
        """
        Get the value of a cell or the value from a merged cell.

        Parameters:
            cell: The cell to retrieve the value from.
            default: The default value to return if the cell is empty or not found.

        Returns:
            Any: The cell's value or the value from a merged cell if applicable, or the default value.
        """
        sheet = cell.parent
        rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
        if len(rng)!=0:
            return sheet.cell(rng[0].min_row, rng[0].min_col).value
        elif cell.value:
            return cell.value
        else: 
            return default
    
    def get_schedule_rows_range(self, ws):
        """
        Get the range of rows containing schedule information in the worksheet.

        Parameters:
            ws: The worksheet to search for schedule rows.

        Returns:
            tuple: A tuple containing the first and last row indices of the schedule rows.
        """
        days_of_week = {
            "Понеділок", 
            "Вівторок", 
            "Середа", 
            "Четвер",
            "П'ятниця",
            "Субота",
            "Неділя"
        }
        

        rows = []
        for cellA, cellB in zip(ws["A"], ws["B"]):
            cellA_val = self.get_cell_val(cellA, default="")
            cellB_val = self.get_cell_val(cellB, default="")

            if  cellB_val and \
                (
                re.match("[0-9]:[0-9][0-9]-[0-9]:[0-9][0-9]"
                        , cellB_val) or 
                re.match("[0-9]:[0-9][0-9]-[0-9][0-9]:[0-9][0-9]"
                        , cellB_val) or
                re.match("[0-9][0-9]:[0-9][0-9]-[0-9]:[0-9][0-9]"
                        , cellB_val) or 
                re.match("[0-9][0-9]:[0-9][0-9]-[0-9][0-9]:[0-9][0-9]"
                        , cellB_val) or
                cellA_val.capitalize() in days_of_week
                ):
                rows.append(cellB.row)
        return rows[0], rows[-1]

    def get_nearest_up_cell_val(self, cell, default = None):
        """
        Get the value of the nearest cell above the given cell.

        Parameters:
            cell: The cell to start searching from.
            default: The default value to return if no value is found.

        Returns:
            Any: The value of the nearest cell above or the default value if not found.
        """
        if self.get_cell_val(cell):
            return self.get_cell_val(cell)
        
        sheet = cell.parent
        for row in range(cell.row-1, 0, -1):
            cell_val = self.get_cell_val(sheet.cell(row = row, column = cell.column))
            if cell_val:
                return cell_val
        
        return default

class DOCXReader(Reader):
    """
    A class for reading schedule data from DOCX files.

    Parameters:
        path (str): The path to the DOCX file.

    """

    def __init__(self, path) -> None:
        """
        Initialize a DOCXReader instance.

        Parameters:
            path (str): The path to the DOCX file.

        """
        super().__init__(path)

        self.WORD_NAMESPACE = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
        self.TEXT = self.WORD_NAMESPACE + 't'
        self.TABLE = self.WORD_NAMESPACE + 'tbl'
        self.ROW = self.WORD_NAMESPACE + 'tr'
        self.CELL = self.WORD_NAMESPACE + 'tc'

    
    def read(self):
        """
        Read data from a DOCX file.

        Returns:
            list: The read schedule data in a structured format.
        """
        
        path = self.path

        with zipfile.ZipFile(path) as docx:
            tree = xml.etree.ElementTree.XML(docx.read('word/document.xml'))
        
        schedule = []

        blank_filler = {   #fill blank
            DAYS_OF_WEEKS: "",
            TIME: "",
            WEEKS: "",
            LECT_HALL: ""
        }
        
        for table_node in tree.iter(self.TABLE):
            
            for row, row_node in enumerate(table_node.iter(self.ROW)):
                if row == 0:
                    continue
                row_list = []

                if len([i for i in row_node.iter(self.CELL)])!=COLS:
                    break

                for col, cell_node in enumerate(row_node.iter(self.CELL)):
                    cell_value = ''.join(node.text for node in cell_node.iter(self.TEXT))
                    if col in blank_filler.keys():
                        if (not cell_value) or cell_value.isspace():
                            cell_value = blank_filler[col]
                        else:
                            blank_filler[col] = cell_value
                        
                    row_list.append(cell_value)

                schedule.append(row_list)
        
        return schedule

class DOCReader(Reader):
    """
    A class for reading schedule data from DOC files.

    Parameters:
        path (str): The path to the DOC file.

    """
    def __init__(self, path) -> None:
        """
        Initialize a DOCReader instance.

        Parameters:
            path (str): The path to the DOC file.

        """
        super().__init__(path)
    
    def read(self):
        """
        Read schedule data from a DOC file.

        Returns:
            list: The read schedule data in a structured format.
        """

        full_path = os.path.abspath(self.path)

        w = wc.Dispatch('Word.Application')

        docx_path = full_path+"x"

        doc = w.Documents.Open(full_path)
        doc.SaveAs(full_path+"x", 16)
        doc.Close()

        w.Quit()

        schedule = DOCXReader(docx_path).read()

        os.remove(docx_path)

        return schedule
    
    
class AbsoluteReader(Reader):
    """
    A class for reading schedule data from files of various formats, determining the format based on file extension.

    Parameters:
        path (str): The path to the data file.

    """
    def __init__(self, path) -> None:
        """
        Initialize an AbsoluteReader instance.

        Parameters:
            path (str): The path to the data file.

        """
        super().__init__(path)

    def read(self):
        """
        Read schedule data from a data file, determining its format based on the file extension.

        Returns:
            list: The read schedule data in a structured format.
        """

        self.path = os.path.abspath(self.path)
        
        filename, file_extension = os.path.splitext(os.path.basename(self.path))
        if file_extension == ".xlsx":
            return XLSXReader(self.path).read()
        elif file_extension == ".docx":
            return DOCXReader(self.path).read()
        elif file_extension == ".doc":
            return DOCReader(self.path).read()
        else:
            raise Exception('extension should be one of [".xlsx", ".docx", ".doc"]')


if __name__ == "__main__":
    from pprint import pprint
    import numpy as np

    #schedule = XLSXReader("./files/Прикладна_математика_БП-4_Осінь_2023–2024.xlsx").read()
    #schedule = DOCXReader("./files/Історія_БП-4_Осінь_2023–2024.docx").read()
    #schedule = DOCReader(r"E:\Misha\GitHub\FIDO_scheduler_test_task_13.10.2023\files\Соціальна_робота_МП-1_Осінь_2023–2024.doc").read()
    schedule = AbsoluteReader("./files/Економіка_БП-1_Осінь_2023–2024 (2).doc").read()

    #print(all(np.array(schedule).flatten()))

    import numpy as np
    import pandas as pd

    my_array = np.array(schedule)

    df = pd.DataFrame(my_array, columns = ['День','Час','Дисципліна, викладач', "Група", "Тижні", "Аудиторія"])

    print(df.to_markdown())
    #pprint(schedule)
