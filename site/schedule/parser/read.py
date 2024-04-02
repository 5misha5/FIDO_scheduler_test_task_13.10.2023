import openpyxl
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import re
from abc import ABC, abstractmethod
import pandas as pd
import numpy as np

import zipfile
import lxml.etree

from win32com import client as wc
import os

import cols

from pprint import pprint


def CellRichText_to_dict(cell: openpyxl.cell.rich_text.CellRichText):
    print(type(cell))
    for i in cell:
        print(i)
        print(i.font)
    rich_cell_text_TextBlocks = map(lambda x: openpyxl.cell.rich_text.TextBlock(x) if isinstance(x,str) else x, cell)
    rich_cell_text_params = map(vars, rich_cell_text_TextBlocks)
    rich_cell_text_params = list(map(lambda p: {**p, "font": vars(p["font"])}, rich_cell_text_params))
    return rich_cell_text_params


def remove_namespace(node):
    """
    Remove namespace prefixes from the given Word XML content.

    Args:
        word_xml (str): The Word XML content.

    Returns:
        str: The Excel XML content with namespace prefixes removed.
    """

    for elem in node.iter():
        # Split the element tag by the namespace prefix and get the last part
        elem.tag = elem.tag.split('}')[-1]

    return node


class Reader(ABC):
    """
    Abstract base class for data readers.

    Parameters:
        path (str): The path to the data file.

    """

    def __init__(self, path) -> None:
        """
        Initialize a Reader instance.

        Parameters:
            path (str): The path to the data file.

        """
        self.path = path

    @abstractmethod
    def read(self):
        """
        Abstract method for reading data from a file.

        Returns:
            list: The read schedule data.
        """
        pass


class XLSXReader(Reader):
    """
    A class for reading schedule data from XLSX files.

    Parameters:
        path (str): The path to the XLSX file.

    """

    def __init__(self, path) -> None:
        """
        Initialize an XLSXReader instance.

        Parameters:
            path (str): The path to the XLSX file.

        """
        super().__init__(path)

    def read(self):
        """
        Read data from an XLSX file.

        Returns:
            list: The read schedule data in a structured format.
        """

        path = self.path

        ws = openpyxl.load_workbook(path, rich_text=True).active

        schedule = []

        rows_range = self.get_schedule_rows_range(ws)

        for row in ws.iter_rows(rows_range[0], rows_range[1]):

            row_data = []

            for col in range(0, 6):
                if col in {cols.DAYS_OF_WEEKS, cols.TIME}:
                    row_data.append(self.get_nearest_up_cell_val(row[col]))
                elif col == cols.COURSE:
                    cell_val = self.get_cell_val(row[col])
                    row_data.append(CellRichText(cell_val if cell_val else "")._opt().as_list())
                else:
                    row_data.append(str(self.get_cell_val(row[col])))

            schedule.append(row_data)

        schedule_df = pd.DataFrame(schedule,
                                   columns=['day_of_week_name',
                                            'time',
                                            'course_lecturer_rich_text',
                                            "group_name",
                                            "weeks",
                                            "auditory_name"])

        return schedule_df

    def get_cell_val(self, cell, default=None):
        """
        Get the value of a cell or the value from a merged cell.

        Parameters:
            cell: The cell to retrieve the value from.

            default: The default value to return if the cell is empty or not found.

        Returns:
            Any: The cell's value or the value from a merged cell if applicable, or the default value.
        """
        sheet = cell.parent
        rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
        if len(rng) != 0:
            return sheet.cell(rng[0].min_row, rng[0].min_col).value
        elif cell.value:
            return cell.value
        else:
            return default

    def get_schedule_rows_range(self, ws):
        """
        Get the range of rows containing schedule information in the worksheet.

        Parameters:
            ws: The worksheet to search for schedule rows.

        Returns:
            tuple: A tuple containing the first and last row indices of the schedule rows.
        """
        days_of_week = {
            "Понеділок",
            "Вівторок",
            "Середа",
            "Четвер",
            "П'ятниця",
            "Субота",
            "Неділя"
        }

        rows = []
        for cellA, cellB in zip(ws["A"], ws["B"]):
            cellA_val = str(self.get_cell_val(cellA, default=""))
            cellB_val = str(self.get_cell_val(cellB, default=""))
            # print(type(cellB_val))
            if cellB_val and \
                    (
                            # if time format
                            re.match('^([01]?[0-9]|2[0-3])[:\.][0-5][0-9]-([01]?[0-9]|2[0-3])[:\.][0-5][0-9]$',

                                     cellB_val) or
                            cellA_val.capitalize() in days_of_week
                    ):
                rows.append(cellB.row)
        return rows[0], rows[-1]

    def get_nearest_up_cell_val(self, cell, default=None):
        """
        Get the value of the nearest cell above the given cell.

        Parameters:
            cell: The cell to start searching from.

            default: The default value to return if no value is found.

        Returns:
            Any: The value of the nearest cell above or the default value if not found.
        """
        if self.get_cell_val(cell):
            return self.get_cell_val(cell)

        sheet = cell.parent
        for row in range(cell.row - 1, 0, -1):
            cell_val = self.get_cell_val(sheet.cell(row=row, column=cell.column))
            if cell_val:
                return cell_val

        return default


class DOCXReader(Reader):
    """
    A class for reading schedule data from DOCX files.

    Parameters:
        path (str): The path to the DOCX file.

    """

    def __init__(self, path) -> None:
        """
        Initialize a DOCXReader instance.

        Parameters:
            path (str): The path to the DOCX file.

        """
        super().__init__(path)

        self.WORD_NAMESPACE = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
        self.TEXT = self.WORD_NAMESPACE + 't'
        self.TABLE = self.WORD_NAMESPACE + 'tbl'
        self.ROW = self.WORD_NAMESPACE + 'tr'
        self.CELL = self.WORD_NAMESPACE + 'tc'

    def read(self):
        """
        Read data from a DOCX file.

        Returns:
            list: The read schedule data in a structured format.
        """

        path = self.path

        with zipfile.ZipFile(path) as docx:
            tree = lxml.etree.XML(docx.read('word/document.xml'))

        schedule = []

        blank_filler = {  # fill blank
            cols.DAYS_OF_WEEKS: "",
            cols.TIME: "",
            cols.WEEKS: "",
            cols.LECT_HALL: ""
        }

        for table_node in tree.iter(self.TABLE):

            for row, row_node in enumerate(table_node.iter(self.ROW)):
                if row == 0:
                    continue
                row_list = []

                if len([i for i in row_node.iter(self.CELL)]) != cols.COLS:
                    break

                for col, cell_node in enumerate(row_node.iter(self.CELL)):

                    if col == cols.COURSE:
                        cell_rich = self.cell_to_CellRichText(cell_node)
                        cell_value = cell_rich._opt().as_list()
                    else:
                        cell_value = "".join(node.text for node in cell_node.iter(self.TEXT))
                        if col in blank_filler.keys():
                            if (not cell_value) or cell_value.isspace():
                                cell_value = blank_filler[col]
                            else:
                                blank_filler[col] = cell_value

                    row_list.append(cell_value)

                schedule.append(row_list)
        schedule_df = pd.DataFrame(schedule,
                                   columns=['day_of_week_name',
                                            'time',
                                            'course_lecturer',
                                            "group_name",
                                            "weeks",
                                            "auditory_name"])

        return schedule_df

    def cell_to_CellRichText(self, cell_node):
        cell_rich = CellRichText()

        for r_element in cell_node.iter(self.WORD_NAMESPACE + "r"):
            t_element = r_element.find(f"{self.WORD_NAMESPACE}t")

            font_attributes = dict()

            if t_element is not None and t_element.text:
                # Extract text content
                text = t_element.text

                # Extract font attributes
                rPr_element = r_element.find(f"{self.WORD_NAMESPACE}rPr")
                for prop in rPr_element.iterchildren():
                    font_attributes[(lxml.etree.QName(prop).localname)] = prop.attrib.values()

            font = InlineFont(rFont=font_attributes["rFonts"][0],
                              sz=int(font_attributes["sz"][0]),
                              i="i" in font_attributes,
                              b="b" in font_attributes
                              )

            cell_rich.append(TextBlock(font, text))

        return cell_rich


class DOCReader(Reader):
    """
    A class for reading schedule data from DOC files.

    Parameters:
        path (str): The path to the DOC file.

    """

    def __init__(self, path) -> None:
        """
        Initialize a DOCReader instance.

        Parameters:
            path (str): The path to the DOC file.

        """
        super().__init__(path)

    def read(self):
        """
        Read schedule data from a DOC file.

        Returns:
            list: The read schedule data in a structured format.
        """

        full_path = os.path.abspath(self.path)

        w = wc.Dispatch('Word.Application')

        docx_path = full_path + "x"

        doc = w.Documents.Open(full_path)
        doc.SaveAs(docx_path, 16)
        doc.Close()

        w.Quit()

        schedule_df = DOCXReader(docx_path).read()

        os.remove(docx_path)

        return schedule_df


class AbsoluteReader(Reader):
    """
    A class for reading schedule data from files of various formats, determining the format based on file extension.

    Parameters:
        path (str): The path to the data file.

    """

    def __init__(self, path) -> None:
        """
        Initialize an AbsoluteReader instance.

        Parameters:
            path (str): The path to the data file.

        """
        super().__init__(path)

    def read(self):
        """
        Read schedule data from a data file, determining its format based on the file extension.

        Returns:
            list: The read schedule data in a structured format.
        """

        self.path = os.path.abspath(self.path)
        print(self.path)

        filename, file_extension = os.path.splitext(os.path.basename(self.path))
        if file_extension == ".xlsx":
            return XLSXReader(self.path).read()
        elif file_extension == ".docx":
            return DOCXReader(self.path).read()
        elif file_extension == ".doc":
            return DOCReader(self.path).read()
        else:
            raise Exception('extension should be one of [".xlsx", ".docx", ".doc"]')


if __name__ == "__main__":
    from pprint import pprint
    import numpy as np

    # schedule = XLSXReader("./files/Прикладна_математика_БП-4_Осінь_2023–2024.xlsx").read()
    # schedule = DOCXReader(r"..\..\..\files/Історія_БП-4_Осінь_2023–2024.docx").read()
    # schedule = DOCReader(r"..\files\Економіка_БП-1_Осінь_2023–2024 (2).doc").read()
    schedule = AbsoluteReader(r"..\..\..\files\Право_БП-1_Осінь_2023–2024.xlsx").read()
    # schedule = AbsoluteReader(r"..\..\..\files\Суспільне_і_приватне_врядування_БП-3_Осінь_2023–2024.xlsx").read()
    # schedule = AbsoluteReader(r"..\..\..\files/Історія_БП-4_Осінь_2023–2024.docx").read()

    # print(all(np.array(schedule).flatten()))

    print(schedule.to_markdown())
    # print(schedule["course_lecturer_rich_text"][0].as_list())
    # print(schedule["course_lecturer_rich_text"][0][1].font)
    # WORD_NAMESPACE = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
    # TEXT = WORD_NAMESPACE + 't'
    # print("".join(node.text for node in lxml.etree.XML(schedule.iloc[0]["course_lecturer"]).iter(TEXT)))
